"""
services/export_service.py

Geração do relatório Excel exportável. Segue o padrão do projeto:
- Datas mantidas como datetime nativo até a última hora, formatadas via
  openpyxl (number_format = "DD/MM/AAAA"), nunca pré-convertidas para
  string — senão o Excel recebe texto e perde ordenação/formatação nativa.
- Cabeçalho estilizado (cor de marca, negrito, texto branco).
- Largura de coluna calculada a partir do conteúdo (auto-fit).
- Painel congelado na primeira linha para facilitar leitura de tabelas
  longas.
- Alinhamento por tipo de dado (texto à esquerda, número à direita,
  status/prioridade centralizados).
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.config import DATE_FORMAT_BR, PALETTE

_HEADER_FILL = PatternFill(
    start_color="0FBF9F", end_color="0FBF9F", fill_type="solid"
)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Calibri", size=10.5)
_CENTER_COLS = {"Status", "Prioridade"}
_THIN_BORDER = Border(bottom=Side(style="thin", color="E2EEEC"))

_EXCEL_DATE_FORMAT = "DD/MM/YYYY"


def _formatar_valor_para_exibicao(value):
    """
    Normaliza valores antes de escrever na célula — evita o problema do
    pandas 3.x onde NaN vira a string "nan" em vez de célula vazia.
    """
    if pd.isna(value):
        return None
    return value


def _autofit_columns(ws: Worksheet, df: pd.DataFrame, min_width: int = 10, max_width: int = 45) -> None:
    """Ajusta a largura de cada coluna com base no maior conteúdo (cabeçalho incluso)."""
    for idx, col in enumerate(df.columns, start=1):
        maior = len(str(col))
        for value in df[col]:
            if pd.isna(value):
                continue
            maior = max(maior, len(str(value)))
        largura = min(max(maior + 3, min_width), max_width)
        ws.column_dimensions[get_column_letter(idx)].width = largura


def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, date_columns: list[str] | None = None) -> None:
    """
    Escreve um DataFrame numa aba já com formatação executiva: cabeçalho
    colorido, datas nativas formatadas, colunas auto-ajustadas, painel
    congelado e alinhamento por tipo de coluna.
    """
    date_columns = date_columns or []
    export_df = df.copy()

    for col in export_df.columns:
        export_df[col] = export_df[col].apply(_formatar_valor_para_exibicao)

    export_df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    # Cabeçalho
    for col_idx, col_name in enumerate(export_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Corpo: fonte, alinhamento, formato de data nativo
    for col_idx, col_name in enumerate(export_df.columns, start=1):
        is_date = col_name in date_columns
        is_numeric = pd.api.types.is_numeric_dtype(df[col_name]) and not is_date
        if col_name in _CENTER_COLS:
            align = Alignment(horizontal="center", vertical="center")
        elif is_numeric:
            align = Alignment(horizontal="right", vertical="center")
        else:
            align = Alignment(horizontal="left", vertical="center")

        for row_idx in range(2, len(export_df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _BODY_FONT
            cell.alignment = align
            cell.border = _THIN_BORDER
            if is_date and cell.value is not None:
                cell.number_format = _EXCEL_DATE_FORMAT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws, export_df)


def build_excel_report(
    tabela_chamados: pd.DataFrame,
    agregado_oficina: pd.DataFrame,
    agregado_categoria: pd.DataFrame,
    date_columns: list[str],
) -> bytes:
    """
    Monta o relatório Excel completo (3 abas) em memória e devolve os
    bytes prontos para o st.download_button.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_sheet(writer, tabela_chamados, "Chamados Filtrados", date_columns=date_columns)
        _write_sheet(writer, agregado_oficina, "Por Oficina")
        _write_sheet(writer, agregado_categoria, "Por Categoria")

    return buffer.getvalue()


def build_excel_report_reposicao(
    tabela_reposicoes: pd.DataFrame,
    agregado_oficina: pd.DataFrame,
    oficinas_pendentes: pd.DataFrame,
    date_columns: list[str],
) -> bytes:
    """
    Monta o relatório Excel do módulo de Reposições (3 abas) em memória e
    devolve os bytes prontos para o st.download_button. Segue o mesmo
    padrão visual de build_excel_report (cabeçalho colorido, auto-fit,
    painel congelado).
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_sheet(writer, tabela_reposicoes, "Reposições Filtradas", date_columns=date_columns)
        _write_sheet(writer, agregado_oficina, "Por Oficina")
        _write_sheet(
            writer,
            oficinas_pendentes,
            "Oficinas Pendentes",
            date_columns=["Solicitação Mais Antiga"],
        )

    return buffer.getvalue()
